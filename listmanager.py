#!/usr/bin/env python
##########################################################################
##                                                                      ##
##   listmanager.py                                                     ##
##                                                                      ##
##   Generate a list of names and emails from an Excel (OOXML .xlsx)    ##
##   workbook that's in a suitable format for LISTSERV's bulk           ##
##   subscription function.                                             ##
##                                                                      ##
##   Author:   Kevin Ernst <ernstki@mail.uc.edu>                        ##
##   Date:     05 November 2016                                         ##
##                                                                      ##
##   References:                                                        ##
##                                                                      ##
##     1. http://pandas.pydata.org/pandas-docs/stable/index.html        ##
##     2. http://openpyxl.readthedocs.io/en/latest/index.html           ##
##                                                                      ##
##   Sample Usage:                                                      ##
##                                                                      ##
##     python workbook.xlsx 2020                                        ##
##                                                                      ##
##   Bugs / Missing Features:                                           ##
##                                                                      ##
##     1. No way to sort output                                         ##
##                                                                      ##
##########################################################################

import os
import sys
import re
import click

from openpyxl import load_workbook
from pandas import DataFrame

class ListManager:
    roster = {}

    def __init__(self, excelfile):
        """
        Open workbook given as argument, look for sheets that end in "20xx"
        and store their contents on a dictionary of DataFrames, with the
        graduating years as (string) keys
        """
        p = re.compile('.*(20\d\d)$')
        wb = load_workbook(excelfile)

        for sheet in wb.sheetnames:
            m = p.match(sheet)
            if not m: continue
            data = wb[sheet].values
            # Skip over any empty rows, then save off the column headers
            cols = next(data)
            while not cols[0]:
                cols = next(data)
            # Store data in dictionary as a Pandas DataFrame
            self.roster[m.group(1)] = DataFrame(data, columns=cols)

        if not self.roster.keys():
            raise RuntimeError("No tabs matching '20xx' found in spreadsheet")

    def print_members(self, year=None):
        """
        Print out the list of members from the all worksheet tabs with a name
        matching "<Something> 20xx" in a format that's suitable for bulk
        import into LISTSERV
        """
        if year:
            years = [str(year)]
        else:
            years = self.roster.keys()

        for year in years:
            df = self.roster[year]
            for p in df[['Name', 'Email']].itertuples():
                if p.Name is None: break
                print("{} <{}>".format(p.Name.title(), p.Email))


@click.command()
@click.argument('excelfile', type=click.Path(exists=True))
@click.argument('year', type=click.INT, required=False)
def read_workbook(excelfile, year=None):
    """Open .xlsx workbook and print names and emails"""
    lm = ListManager(excelfile)
    lm.print_members(year)

if __name__ == '__main__':
    read_workbook()
